"""Inspect Project 5's M64 cell — what makes it $281k?"""
from openpyxl import load_workbook
from pathlib import Path

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wbv = load_workbook(SRC, data_only=True)
wb = load_workbook(SRC, data_only=False)

for tab in ["Project 5", "Project 6", "Project 2 - 84 SBR"]:
    ws = wbv[tab]
    wsf = wb[tab]
    print(f"\n=== {tab} ===")
    for r in range(60, 75):
        val = ws.cell(r, 13).value
        formula = wsf.cell(r, 13).value
        b = ws.cell(r, 2).value
        if val is not None or formula or b:
            print(f"  Row {r} B={b}  M{r} value={val}  formula={formula}")
