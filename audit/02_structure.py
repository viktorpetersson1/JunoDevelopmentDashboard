"""Phase 1 — dump row labels (column A) and column headers (row 1-12) for every sheet."""
from openpyxl import load_workbook
from pathlib import Path
import json

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
OUT = Path(r"C:\Dev\juno-financial-dashboard\audit")

wb = load_workbook(SRC, data_only=False)
wb_v = load_workbook(SRC, data_only=True)

result = {}
for name in wb.sheetnames:
    ws = wb[name]
    wsv = wb_v[name]
    row_labels = []
    for r in range(1, min(ws.max_row + 1, 200)):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is not None or b is not None:
            row_labels.append({"row": r, "A": str(a)[:120] if a is not None else None, "B": str(b)[:80] if b is not None else None})
    headers_top = []
    for r in range(1, min(ws.max_row + 1, 13)):
        row = []
        for c in range(1, min(ws.max_column + 1, 70)):
            v = ws.cell(r, c).value
            if v is not None:
                row.append({"c": c, "v": str(v)[:60]})
        if row:
            headers_top.append({"row": r, "cells": row})
    result[name] = {
        "state": ws.sheet_state,
        "dims": f"{ws.max_row}x{ws.max_column}",
        "row_labels_A_B": row_labels,
        "top_rows": headers_top,
    }

(OUT / "02_structure.json").write_text(json.dumps(result, indent=2, default=str))
print("Saved 02_structure.json")
