"""Roll up the 84 SBR takeoff by CSI category code."""
from openpyxl import load_workbook
from pathlib import Path
import json, re

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wbv = load_workbook(SRC, data_only=True)
ws = wbv["Construction costs 84SB"]

# Map CSI prefixes to human names
csi_names = {
    "000": "Production of panels (Kingshaus)",
    "00":  "Procurement & contracting",
    "01":  "General requirements",
    "02":  "Existing conditions",
    "03":  "Concrete & foundation",
    "04":  "Masonry",
    "05":  "Metals",
    "06":  "Carpentry & wood",
    "07":  "Thermal & moisture protection",
    "08":  "Doors & windows",
    "09":  "Finishes",
    "10":  "Specialties",
    "11":  "Equipment",
    "12":  "Furnishings",
    "13":  "Special construction",
    "14":  "Conveying systems",
    "15":  "Mechanical",
    "16":  "Electrical",
    "21":  "Fire suppression",
    "22":  "Plumbing",
    "23":  "HVAC",
    "26":  "Electrical (CSI 2014)",
    "31":  "Earthwork",
    "32":  "Exterior improvements",
    "33":  "Utilities",
    "34":  "Misc",
    "99":  "Build fee",
}

categories = {}
for r in range(3, ws.max_row + 1):
    code = ws.cell(r, 1).value
    name = ws.cell(r, 2).value
    final = ws.cell(r, 8).value  # column H = final estimate
    if not isinstance(code, str) or not isinstance(final, (int, float)):
        continue
    # Extract prefix
    m = re.match(r"^(\d{2,3})", code.strip())
    if not m:
        continue
    prefix = m.group(1)
    bucket = csi_names.get(prefix, f"Other ({prefix})")
    if bucket not in categories:
        categories[bucket] = {"total": 0, "lines": []}
    categories[bucket]["total"] += final
    categories[bucket]["lines"].append({"code": code.strip(), "name": str(name).strip(), "amount": round(final, 2)})

# Sort buckets by total descending
sorted_categories = sorted(categories.items(), key=lambda kv: -kv[1]["total"])

out = {"project": "84 SBR (Project 2)", "source": "Construction costs 84SB tab", "snapshot": "2026-05-10", "categories": []}
for name, data in sorted_categories:
    out["categories"].append({"name": name, "total": round(data["total"], 2), "lines": data["lines"]})

out_path = Path(r"C:\Dev\juno-financial-dashboard\audit\84sbr_takeoff.json")
out_path.write_text(json.dumps(out, indent=2))
print(f"Saved {out_path}")
print(f"\nTotal categories: {len(sorted_categories)}")
print(f"\nTop 10 by total cost:")
for name, data in sorted_categories[:10]:
    print(f"  ${data['total']:>10,.0f}  {name} ({len(data['lines'])} lines)")
total = sum(d["total"] for _, d in sorted_categories)
print(f"\nGrand total: ${total:,.0f}")
