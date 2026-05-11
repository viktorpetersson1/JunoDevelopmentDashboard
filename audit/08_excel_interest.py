"""Extract Excel's actual interest values per project + understand cumulative basis."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wbv = load_workbook(SRC, data_only=True)
wb = load_workbook(SRC, data_only=False)

# All projects — sum row 64 (interest) and row 71 (total financing) across the full row
projects = ["Project 2 - 84 SBR", "Project 3 - TBC", "Project 4 - Hands Creek"] + [f"Project {i}" for i in range(5, 12)]
print(f"{'Project':<25} {'Interest (r64)':>18} {'Financing (r71)':>18} {'M14 build':>14} {'M27 sale':>14}")
total_int = 0
total_fin = 0
for tab in projects:
    if tab not in wbv.sheetnames: continue
    ws = wbv[tab]
    r64 = sum((v if isinstance(v, (int, float)) else 0) for v in (ws.cell(64, c).value for c in range(3, ws.max_column + 1)))
    r71 = sum((v if isinstance(v, (int, float)) else 0) for v in (ws.cell(71, c).value for c in range(3, ws.max_column + 1)))
    m14 = ws.cell(14, 13).value
    m27 = ws.cell(27, 13).value
    print(f"{tab:<25} {r64:>18,.0f} {r71:>18,.0f} {m14 if m14 else 0:>14,.0f} {m27 if m27 else 0:>14,.0f}")
    total_int += r64
    total_fin += r71

print(f"{'TOTAL':<25} {total_int:>18,.0f} {total_fin:>18,.0f}")

# Pick one project (5) and dump the months where row 64 is non-zero
ws = wbv["Project 5"]
print(f"\n=== Project 5 — non-zero interest months ===")
for c in range(3, ws.max_column + 1):
    v = ws.cell(64, c).value
    if isinstance(v, (int, float)) and v != 0:
        eq = ws.cell(85, c).value
        hd = ws.cell(88, c).value
        print(f"  {get_column_letter(c)}64 (month col {c}) interest={v:,.0f}  cum_eq(85)={eq if eq else 0:,.0f}  hard(88)={hd if hd else 0:,.0f}")
