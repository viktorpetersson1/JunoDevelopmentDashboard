"""Phase 1 inspection — read-only audit of the Juno financial model."""
from openpyxl import load_workbook
from pathlib import Path
import json

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
OUT = Path(r"C:\Dev\juno-financial-dashboard\audit")

# Pass 1 — formulas (keep_vba=False, data_only=False)
wb = load_workbook(SRC, data_only=False, read_only=False)
# Pass 2 — values (we'll load separately for value snapshot)
wb_vals = load_workbook(SRC, data_only=True, read_only=False)

summary = {
    "file": str(SRC),
    "sheets": [],
    "defined_names": [],
    "external_links": [],
}

# Defined names (named ranges)
try:
    for dn in wb.defined_names:
        ref = wb.defined_names[dn]
        summary["defined_names"].append({"name": dn, "value": str(ref.value) if hasattr(ref, "value") else str(ref)})
except Exception as e:
    summary["defined_names_error"] = str(e)

# External links
try:
    if hasattr(wb, "_external_links"):
        for el in wb._external_links:
            summary["external_links"].append(str(el))
except Exception as e:
    summary["external_links_error"] = str(e)

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    ws_v = wb_vals[ws_name]
    info = {
        "name": ws_name,
        "state": ws.sheet_state,  # visible / hidden / veryHidden
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_cells": len(ws.merged_cells.ranges),
        "freeze_panes": ws.freeze_panes,
        "tab_color": str(ws.sheet_properties.tabColor) if ws.sheet_properties.tabColor else None,
        "data_validations": len(ws.data_validations.dataValidation) if ws.data_validations else 0,
    }
    summary["sheets"].append(info)

OUT.mkdir(exist_ok=True, parents=True)
(OUT / "01_workbook_overview.json").write_text(json.dumps(summary, indent=2, default=str))
print(json.dumps(summary, indent=2, default=str))
