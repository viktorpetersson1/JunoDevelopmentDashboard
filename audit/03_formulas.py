"""Phase 1 — extract all formulas, classify hardcodes/errors/cross-sheet links."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import re, json, collections

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
OUT = Path(r"C:\Dev\juno-financial-dashboard\audit")

wb = load_workbook(SRC, data_only=False)
wb_v = load_workbook(SRC, data_only=True)

ERR_RE = re.compile(r"#(REF|DIV/0|VALUE|NAME|NUM|N/A|NULL)!")
EXT_RE = re.compile(r"\[\d+\]")  # external link refs like [1]Sheet!A1
SHEETREF_RE = re.compile(r"'([^']+)'!|([A-Za-z_][A-Za-z0-9_]*)!", re.IGNORECASE)
NUM_RE = re.compile(r"(?<![A-Z0-9_!:'\"$])(-?\d+(?:\.\d+)?)(?![A-Z0-9_:])", re.IGNORECASE)

report = {
    "totals": {"formulas": 0, "values": 0, "blanks": 0, "errors": 0, "ext_links": 0, "hardcoded_in_formula": 0},
    "by_sheet": {},
    "errors": [],
    "external_link_cells": [],
    "formulas_with_numeric_literals": [],
    "cross_sheet_edges": collections.Counter(),
}

for name in wb.sheetnames:
    ws = wb[name]
    wsv = wb_v[name]
    sheet_stats = {"formulas": 0, "values": 0, "blanks": 0, "errors": 0, "ext_links": 0}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                sheet_stats["blanks"] += 1
                continue
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                sheet_stats["formulas"] += 1
                report["totals"]["formulas"] += 1
                # cross-sheet refs
                for m in SHEETREF_RE.finditer(v):
                    s = m.group(1) or m.group(2)
                    if s and s != name:
                        report["cross_sheet_edges"][(name, s)] += 1
                # external link?
                if EXT_RE.search(v):
                    sheet_stats["ext_links"] += 1
                    report["totals"]["ext_links"] += 1
                    report["external_link_cells"].append({"sheet": name, "cell": cell.coordinate, "formula": v[:200]})
                # numeric literal inside a formula = hardcode
                literals = []
                for m in NUM_RE.finditer(v):
                    n = m.group(1)
                    # skip simple multipliers we expect like /1000, *12, 0, 1 (common in indexing/calendar)
                    try:
                        nv = float(n)
                        if abs(nv) <= 1 or abs(nv) in (12, 100, 1000, 10000):
                            continue
                        literals.append(n)
                    except ValueError:
                        pass
                if literals:
                    report["totals"]["hardcoded_in_formula"] += 1
                    report["formulas_with_numeric_literals"].append({"sheet": name, "cell": cell.coordinate, "formula": v[:200], "literals": literals})
                # error in literal formula? Most errors arrive via data_only
            else:
                sheet_stats["values"] += 1
                report["totals"]["values"] += 1
            # check evaluated value for #REF! etc.
            try:
                ev = wsv[cell.coordinate].value
                if isinstance(ev, str) and ERR_RE.search(ev):
                    sheet_stats["errors"] += 1
                    report["totals"]["errors"] += 1
                    report["errors"].append({"sheet": name, "cell": cell.coordinate, "value": ev, "formula": v if isinstance(v, str) and v.startswith("=") else None})
            except Exception:
                pass
    report["by_sheet"][name] = sheet_stats

# serialize counter
report["cross_sheet_edges"] = [{"from": k[0], "to": k[1], "count": v} for k, v in report["cross_sheet_edges"].most_common()]

# trim
report["formulas_with_numeric_literals_sample"] = report["formulas_with_numeric_literals"][:50]
report["formulas_with_numeric_literals_count"] = len(report["formulas_with_numeric_literals"])
del report["formulas_with_numeric_literals"]

(OUT / "03_formulas_report.json").write_text(json.dumps(report, indent=2, default=str))
print(json.dumps({"totals": report["totals"], "by_sheet": report["by_sheet"], "errors_n": len(report["errors"]), "ext_link_cells_n": len(report["external_link_cells"])}, indent=2))
