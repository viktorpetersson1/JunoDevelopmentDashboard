"""Sanity check: list labels in column B for each project tab, rows 1-40, side-by-side."""
from openpyxl import load_workbook
from pathlib import Path
import json

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
OUT = Path(r"C:\Dev\juno-financial-dashboard\audit")

wb = load_workbook(SRC, data_only=False)
wb_v = load_workbook(SRC, data_only=True)
project_tabs = [n for n in wb.sheetnames if n.startswith("Project ") and "x" not in n.split()[-1].lower()] + ["6 GC"]
# Actually include all
project_tabs = ["Project 2 - 84 SBR", "Project 3 - TBC", "Project 4 - Hands Creek"] + [f"Project {i}" for i in range(5, 12)] + ["6 GC"]

rows = {}
for r in range(1, 115):
    row = {}
    for tab in project_tabs:
        if tab not in wb.sheetnames:
            continue
        ws = wb[tab]
        c = ws.cell(r, 2).value  # column B label
        if c is not None:
            row[tab] = str(c)[:60]
    if row:
        rows[r] = row

# Also extract input column (col M = 13) values, rows 1-40
inputs = {}
for r in range(1, 40):
    row = {}
    for tab in project_tabs:
        if tab not in wb.sheetnames: continue
        ws = wb[tab]
        wsv = wb_v[tab]
        label_b = ws.cell(r, 2).value
        label_c = ws.cell(r, 3).value
        m_val = wsv.cell(r, 13).value  # evaluated value
        m_formula = ws.cell(r, 13).value if isinstance(ws.cell(r, 13).value, str) and str(ws.cell(r, 13).value).startswith("=") else None
        if label_b or m_val is not None:
            row[tab] = {"B": str(label_b)[:50] if label_b else None, "C": str(label_c)[:50] if label_c else None, "M_val": str(m_val)[:60] if m_val is not None else None, "M_formula": m_formula}
    if row:
        inputs[r] = row

(OUT / "05_project_rowmap.json").write_text(json.dumps({"labels_B": rows, "inputs_col_M": inputs}, indent=2, default=str))
print("Saved 05_project_rowmap.json — rows:", len(rows), "input rows:", len(inputs))
