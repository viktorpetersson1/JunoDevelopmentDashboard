"""Check Juno Forecast row 56 (Equity) and Project 2 row 108 formulas."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wb = load_workbook(SRC, data_only=False)
wbv = load_workbook(SRC, data_only=True)

ws = wb["Juno Forecast"]
print("=== Juno Forecast row 56 (Equity) sample formulas ===")
for c in [5, 6, 7, 11, 18]:
    f = ws.cell(56, c).value
    v = wbv["Juno Forecast"].cell(56, c).value
    print(f"  {get_column_letter(c)}56: formula={f}")
    print(f"           value={v}")

print("\n=== Project 2 row 108 sample formulas ===")
p2 = wb["Project 2 - 84 SBR"]
p2v = wbv["Project 2 - 84 SBR"]
for c in [15, 17, 18, 20, 25, 30]:  # O, Q, R, T, Y, AD
    f = p2.cell(108, c).value
    v = p2v.cell(108, c).value
    print(f"  {get_column_letter(c)}108: formula={f}  value={v}")

print("\n=== Project 2 row 109 (next row, might be paired) ===")
for c in [15, 17, 18, 20, 25, 30]:
    f = p2.cell(109, c).value
    v = p2v.cell(109, c).value
    print(f"  {get_column_letter(c)}109: formula={f}  value={v}")

print("\n=== Project 2 row 110 (might be the debt row) ===")
for c in [15, 17, 18, 20, 25, 30]:
    f = p2.cell(110, c).value
    v = p2v.cell(110, c).value
    print(f"  {get_column_letter(c)}110: formula={f}  value={v}")

print("\n=== Project 2 row 73 (might be the source) ===")
for c in [15, 17, 18, 20, 25, 30]:
    f = p2.cell(73, c).value
    v = p2v.cell(73, c).value
    print(f"  {get_column_letter(c)}73: formula={f}  value={v}")

print("\n=== Project 2 col Q (Apr 26) rows 38-95 ===")
for r in range(38, 95):
    f = p2.cell(r, 17).value
    v = p2v.cell(r, 17).value
    b = p2.cell(r, 2).value or ""
    if f or (v is not None and v != 0):
        print(f"  Q{r} {str(b)[:30]:<30} formula={str(f)[:50] if f else '':<50} value={v}")
