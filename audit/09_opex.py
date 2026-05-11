"""Extract monthly OPEX from Juno Opex Forecast tab — categorized totals per month."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import json

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wb = load_workbook(SRC, data_only=True)
ws = wb["Juno Opex Forecast"]

# Find the date row (usually row 4 or 5)
date_row = None
for r in range(1, 10):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if hasattr(v, "year") and 2025 <= v.year <= 2030:
            date_row = r
            break
    if date_row: break

print(f"Date row: {date_row}")
if date_row:
    dates = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(date_row, c).value
        if hasattr(v, "year"):
            dates.append((c, v.strftime("%Y-%m")))
    print(f"Dates: {dates[:5]} ... {dates[-3:]}")

# Sum each month: look for the "Total Expense" or "Total Operating Expenses" row
print("\n=== Sheet rows with text labels ===")
for r in range(1, min(ws.max_row + 1, 70)):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    if a and isinstance(a, str) and ("Total" in a or "Net" in a or "Expense" in a or "Income" in a):
        last_col_val = ws.cell(r, ws.max_column).value
        print(f"  R{r} A={a[:50]}  last_col_val={last_col_val}")
    elif b and isinstance(b, str) and ("Total" in b or "Net" in b or "Expense" in b):
        print(f"  R{r} B={b[:50]}")

# Try to find Total Expense row and extract monthly values
monthly = {}
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    if a and isinstance(a, str) and "Total Expense" in a:
        print(f"\nTotal Expense row = {r}")
        for c, ym in dates:
            v = ws.cell(r, c).value
            if isinstance(v, (int, float)):
                monthly[ym] = v
        break

print(f"\nMonthly opex {len(monthly)} entries")
for k in list(monthly.keys())[:6] + (["..."] if len(monthly) > 12 else []) + list(monthly.keys())[-3:]:
    print(f"  {k} = {monthly.get(k, '...')}")

print(f"\nTotal across all months: {sum(monthly.values()):,.0f}")
out = Path(r"C:\Dev\juno-financial-dashboard\audit\09_opex.json")
out.write_text(json.dumps(monthly, indent=2, default=str))
print(f"Saved {out}")
