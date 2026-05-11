"""Phase 1 — drill into errors, external links, project tab anatomy, summary tab logic."""
from openpyxl import load_workbook
from pathlib import Path
import json, re

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
OUT = Path(r"C:\Dev\juno-financial-dashboard\audit")

wb = load_workbook(SRC, data_only=False)
wb_v = load_workbook(SRC, data_only=True)

out = {"financing_84sb_errors": [], "external_links_detail": [], "project_2_anatomy": {}, "summary_logic": [], "kpc_equity_flow": []}

# Financing 84SB errors
ws = wb["Financing 84SB"]; wsv = wb_v["Financing 84SB"]
for row in ws.iter_rows():
    for cell in row:
        if cell.value is None: continue
        ev = wsv[cell.coordinate].value
        if isinstance(ev, str) and "#" in ev and "!" in ev:
            out["financing_84sb_errors"].append({"cell": cell.coordinate, "label_A": ws.cell(cell.row, 1).value, "label_B": ws.cell(cell.row, 2).value, "formula": cell.value, "evaluated": ev})

# External links — look up actual references
for name in ["Project 3x", "Project 2 - 84 SBR", "6 GC", "Juno Opex Forecast"]:
    if name not in wb.sheetnames: continue
    ws = wb[name]
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "[" in cell.value and "]" in cell.value and cell.value.startswith("="):
                out["external_links_detail"].append({"sheet": name, "cell": cell.coordinate, "row_label": ws.cell(cell.row, 1).value or ws.cell(cell.row, 2).value, "formula": cell.value[:300]})

# Project 2 - 84 SBR full anatomy: row labels + sample formulas across columns
ws = wb["Project 2 - 84 SBR"]
anatomy = []
for r in range(1, ws.max_row + 1):
    label_a = ws.cell(r, 1).value
    label_b = ws.cell(r, 2).value
    # Sample C, K, O, S columns to see formula pattern across time
    samples = []
    for c in [3, 5, 10, 15, 20, 25, 30, 35]:
        if c <= ws.max_column:
            v = ws.cell(r, c).value
            if v is not None:
                samples.append({"col": c, "val": str(v)[:120]})
    if label_a or label_b or samples:
        anatomy.append({"row": r, "A": str(label_a)[:120] if label_a else None, "B": str(label_b)[:80] if label_b else None, "samples": samples})
out["project_2_anatomy"] = anatomy

# Summary tab — full
ws = wb["Summary"]
for r in range(1, ws.max_row + 1):
    row_data = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None:
            row_data.append({"col": c, "val": str(v)[:200]})
    if row_data:
        out["summary_logic"].append({"row": r, "cells": row_data})

# KPC Equity Flow — full
ws = wb["KPC Equity Flow"]
for r in range(1, ws.max_row + 1):
    row_data = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(r, c).value
        if v is not None:
            row_data.append({"col": c, "val": str(v)[:200]})
    if row_data:
        out["kpc_equity_flow"].append({"row": r, "cells": row_data})

(OUT / "04_drill.json").write_text(json.dumps(out, indent=2, default=str))
print("Saved 04_drill.json — sizes:", {k: len(v) for k, v in out.items()})
