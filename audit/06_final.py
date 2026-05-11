"""Final pass — pull key snapshot values, Juno Forecast aggregation, overhead totals."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
import json

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
OUT = Path(r"C:\Dev\juno-financial-dashboard\audit")

wb = load_workbook(SRC, data_only=False)
wbv = load_workbook(SRC, data_only=True)

result = {}

# Juno Forecast — key headline values per year and overall
ws = wbv["Juno Forecast"]
# rows: 8 sales, 10 land, 11 construction, 12 kingshaus, 13 total dev, 15 overheads, 16 financing, 17 profit
# annual columns: BA=53 FY26, BB=54 FY27, BC=55 FY28, BD=56 FY29 (and maybe BE=57 FY30, BG=59 Total)
labels_by_row = {8:"Sales", 10:"Land cost", 11:"Construction costs", 12:"Kingshaus", 13:"Total dev costs", 15:"Overheads", 16:"Financing costs", 17:"Profit before tax", 81:"Equity requirement (monthly)", 83:"Max equity", 67:"Cumulative debt"}
year_cols = {"FY26": 53, "FY27": 54, "FY28": 55, "FY29": 56}
totals_by_metric = {}
for r, lbl in labels_by_row.items():
    row_vals = {}
    for y, c in year_cols.items():
        v = ws.cell(r, c).value
        row_vals[y] = v
    totals_by_metric[lbl] = row_vals
result["juno_forecast_annual"] = totals_by_metric

# Aggregation pattern check: confirm row 8 formula uses sum of project!row39
ws_f = wb["Juno Forecast"]
sample_formulas = {}
for r, lbl in labels_by_row.items():
    sample_formulas[lbl] = ws_f.cell(r, 5).value  # column E = month 3 (Mar 2026)
result["juno_forecast_e_column_formulas"] = sample_formulas

# Project headline figures (col M outputs)
project_tabs = ["Project 2 - 84 SBR", "Project 3 - TBC", "Project 4 - Hands Creek"] + [f"Project {i}" for i in range(5, 12)]
project_kpis = {}
metric_rows = {"Villa sqft (M10)": 10, "Land cost (M9)": 9, "Build $/sqft (M13)": 13, "Total build cost (M14)": 14, "Interest rate (M16)": 16, "LTC (M17)": 17, "Margin (M22)": 22, "Sale $/sqft (M23)": 23, "Sale price (M27)": 27, "Total profit (M28)": 28}
for tab in project_tabs:
    if tab not in wbv.sheetnames: continue
    wsv = wbv[tab]
    proj = {}
    for m, r in metric_rows.items():
        proj[m] = wsv.cell(r, 13).value
    project_kpis[tab] = proj
result["project_kpis"] = project_kpis

# Juno Opex Forecast — extract annual totals
ws = wbv["Juno Opex Forecast"]
opex_rows = []
for r in range(1, ws.max_row + 1):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    # last column likely total
    last_col_val = ws.cell(r, ws.max_column).value
    if a or b or c or last_col_val is not None:
        opex_rows.append({"row": r, "A": str(a)[:60] if a else None, "B": str(b)[:60] if b else None, "C": str(c)[:60] if c else None, "total_last_col": last_col_val})
result["juno_opex_rows"] = opex_rows[:80]

# Summary D91, D96 - the global assumption cells referenced from projects
wsv = wbv["Summary"]
wsf = wb["Summary"]
result["summary_drivers"] = {
    "D91 (referenced by Project M13)": {"value": wsv["D91"].value, "formula": wsf["D91"].value},
    "D96 (referenced by Project M22)": {"value": wsv["D96"].value, "formula": wsf["D96"].value},
    "F102:O111 (project start/sale dates)": [[wsv.cell(r, c).value for c in range(6, 16)] for r in range(102, 112)],
}

# Save
(OUT / "06_final.json").write_text(json.dumps(result, indent=2, default=str))
print(json.dumps(result, indent=2, default=str)[:8000])
