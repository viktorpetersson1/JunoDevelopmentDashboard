"""Extract the detailed 84 SBR construction cost takeoff from Excel."""
from openpyxl import load_workbook
from pathlib import Path
import json

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wbv = load_workbook(SRC, data_only=True)
wb = load_workbook(SRC, data_only=False)

ws = wbv["Construction costs 84SB"]
wsf = wb["Construction costs 84SB"]

print("=== Construction costs 84SB structure ===")
print(f"Dims: {ws.max_row} x {ws.max_column}")

# Find row labels in col A or B
print("\n=== Rows 1-40 (top of sheet) ===")
for r in range(1, 40):
    a = ws.cell(r, 1).value
    b = ws.cell(r, 2).value
    c = ws.cell(r, 3).value
    h = ws.cell(r, 8).value  # column H = potential total
    if a or b or c or h:
        print(f"R{r}: A={str(a)[:30] if a else ''}  B={str(b)[:40] if b else ''}  C={str(c)[:30] if c else ''}  H={h}")

# Find the H122 value referenced from Project 2!M13 (Build $/sqft)
print(f"\nH122 (Build $/sqft total): {ws.cell(122, 8).value}")
print(f"Formula: {wsf.cell(122, 8).value}")

# Find the total cost rows
print("\n=== Rows 115-130 (bottom of sheet) ===")
for r in range(115, 130):
    cells = [ws.cell(r, c).value for c in range(1, 12)]
    if any(c is not None for c in cells):
        print(f"R{r}: {[str(c)[:25] if c else '' for c in cells]}")
