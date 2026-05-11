"""Diagnose Excel's peak equity calculation: row 83 'Max equity' on Juno Forecast."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wbv = load_workbook(SRC, data_only=True)
wb = load_workbook(SRC, data_only=False)

ws = wbv["Juno Forecast"]
wsf = wb["Juno Forecast"]

print("=== Juno Forecast rows 76-83 (equity/cash schedule) ===")
for r in range(74, 87):
    b = ws.cell(r, 2).value
    if not b: continue
    print(f"\nR{r} B = {b}")
    # Show C, E, K, R, X, AC, AY columns (mar-26, sep-26, jan-27, jun-27, etc.)
    for c in [3, 5, 11, 18, 24, 30, 36, 42, 51]:
        v = ws.cell(r, c).value
        f = wsf.cell(r, c).value
        col = get_column_letter(c)
        if v is not None or (isinstance(f, str) and f.startswith("=")):
            print(f"  {col}{r}: value={v if isinstance(v, (int,float)) else v}  formula={f}")

# Row 83 specifically — what formula?
print(f"\n=== Row 83 'Max equity' formulas in first 5 cols ===")
for c in range(3, 8):
    print(f"  {get_column_letter(c)}83: value={ws.cell(83, c).value}  formula={wsf.cell(83, c).value}")

# C83 = the cell referenced by Summary!D6 = Peak equity required
print(f"\nC83 (the peak equity reported on Summary) = {ws.cell(83, 3).value}")
print(f"Formula at C83 = {wsf.cell(83, 3).value}")

# Row 82 — cumulative equity. Look at the trajectory
print("\n=== Row 82 'Cumulative equity' across all months ===")
peak = 0; peak_col = None
for c in range(3, ws.max_column + 1):
    v = ws.cell(82, c).value
    if isinstance(v, (int, float)):
        if v > peak:
            peak = v
            peak_col = c
print(f"Peak cumulative equity (row 82): {peak:,.2f} at column {get_column_letter(peak_col) if peak_col else 'n/a'}")

# Row 81 — monthly equity requirement
print("\n=== Row 81 'Equity requirement' — non-zero months ===")
total = 0
for c in range(3, ws.max_column + 1):
    v = ws.cell(81, c).value
    if isinstance(v, (int, float)) and v > 0:
        total += v
        if c % 6 == 0:  # every 6th non-zero for brevity
            print(f"  {get_column_letter(c)}81: {v:,.0f}")
print(f"Sum row 81 = {total:,.2f}")

# Row 80 — cash before equity
print("\n=== Row 80 'Cash before equity' (negative = need equity) ===")
for c in [3, 5, 11, 18, 24, 30, 36, 42, 51]:
    v = ws.cell(80, c).value
    f = wsf.cell(80, c).value
    print(f"  {get_column_letter(c)}80: value={v}  formula={f}")
