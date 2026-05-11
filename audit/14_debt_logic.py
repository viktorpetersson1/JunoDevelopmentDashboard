"""Diagnose Excel's debt drawdown + early cash flow."""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path

SRC = Path(r"C:\Dev\juno-financial-dashboard\audit\_source_readonly.xlsx")
wbv = load_workbook(SRC, data_only=True)
wb = load_workbook(SRC, data_only=False)

ws = wbv["Juno Forecast"]
wsf = wb["Juno Forecast"]

# Trace cash flow Mar 26 (E=5) through Apr 27 (R=18)
print("=== Cash flow trace Mar 26 (E) -> Apr 27 (R) ===")
print(f"{'Row':<4} {'Label':<28}", end="")
for c in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
    print(f"{get_column_letter(c):>11}", end="")
print()

interesting_rows = [
    (8,  "Sales"),
    (10, "Land cost"),
    (11, "Construction"),
    (12, "Kingshaus"),
    (13, "Total dev cost"),
    (15, "Overheads"),
    (16, "Financing cost"),
    (17, "Profit before tax"),
    (43, "Sale of villa"),
    (44, "Loan repayment"),
    (46, "Land costs"),
    (47, "Construction costs"),
    (48, "Superstructure (King)"),
    (49, "Financing costs"),
    (50, "Overheads"),
    (51, "Total construction"),
    (55, "Debt"),
    (56, "Equity"),
    (57, "Financing"),
    (59, "Net cash flow"),
    (67, "Cumulative debt"),
    (68, "Cumulative equity"),
    (76, "Opening cash"),
    (77, "Net cash before fin"),
    (78, "Closing cash"),
    (80, "Cash before equity"),
    (81, "Equity requirement"),
    (82, "Cumilative equity"),
]
for r, label in interesting_rows:
    print(f"{r:<4} {label:<28}", end="")
    for c in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
        v = ws.cell(r, c).value
        if isinstance(v, (int, float)):
            print(f"{v:>11,.0f}", end="")
        else:
            print(f"{'':>11}", end="")
    print()

print("\n=== Row 55 'Debt' formula (sample cells) ===")
for c in [5, 8, 11, 14, 18]:
    f = wsf.cell(55, c).value
    print(f"  {get_column_letter(c)}55: formula={f}")
